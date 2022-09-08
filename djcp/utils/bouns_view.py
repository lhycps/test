import datetime
import os
import time
from functools import reduce

import openpyxl
from django.utils.http import urlquote

from djcp import utils
import logging
from django.core.files.base import ContentFile
from django.db.models import Q
from django.http import JsonResponse, StreamingHttpResponse, HttpResponse
from django.utils.encoding import escape_uri_path
from django.views.decorators.csrf import csrf_exempt


from crm import settings
from djcp.utils import util

from djcp.models import ProInfo, BonusUnit, BonusSystem

log = logging.getLogger('mydjango')  # 这里的mydjango是settings中loggers里面对应的名字


def insertVal(request):
    '''
    点击select输入框插入数据到对应的input输入框中
    :param request:
    :return:
    '''
    res = {'code': '200'}
    bid = request.GET.get('bid')
    prolist = ProInfo.objects.filter(djcp__customer=bid).values('sys_name', 'level', 'djcp__prodate__agreen',
                                                                'djcp__customer__unit_name', 'pm')

    prolist = list(prolist)

    return JsonResponse(prolist, safe=False)


@csrf_exempt
def bonus_add(request):
    '''
    添加奖金
    :param request:
    :return:
    '''
    user = request.user
    res = {'user': None, 'msg': None}
    hh = request.FILES.get('bonus_pdf', '')
    input_data = request.POST.get('inputdata')
    newdata = util.formdata_to_dict(input_data)
    _dic_start = util.dic_slice(newdata, 'nature', 'sale')
    _dic_end = util.dic_slice(newdata, 'MAX_FILE_SIZE', 'remarks')
    dd = {}
    dd.update(_dic_start)
    dd.update(_dic_end)
    ss = {}
    for i, j in newdata.items():
        if not i in dd:
            ss[i] = j
    new_list = util.dict_chunk(ss, 3)

    try:
        if dd['bonus_pro'] == '':  # 如果绑定项目的value值为空说明走的逻辑是自动添加逻辑
            util.Autonomous_addition(new_list, dd, user, hh)
        else:  # 走的是绑定系统的逻辑
            util.bindProAdd(new_list, dd, user, hh)

        res['msg'] = '新建成功'
    except:
        res['error'] = '错误，可能数据库中已存在该数据或者您提交的数据类型为空'

    return JsonResponse(res, safe=False)


def bonus_edit(request):
    # 点击编辑按钮弹出编辑模态对话框并将数据插入到对应的模态对话框中
    res = {'code': "200"}
    bonus_editid = request.GET.get('bonus_editid')
    bonus = BonusUnit.objects.filter(id=bonus_editid).values('nature', 'unit', 'bonus', 'pm', 'bonus_pro',
                                                             'sale', 'contract_pdf',
                                                             'remarks', 'bonussystem__system',
                                                             'bonussystem__level', 'bonussystem__completion_time')

    if bonus:
        bonus = list(bonus)
        bonus[0]['bonus'] = util.decrypt(bonus[0]['bonus'])
        print("bonus", bonus)

        res['msg'] = bonus
    else:
        res['error'] = '无此奖金相关信息'
    return JsonResponse(res, safe=False)


@csrf_exempt
def bonus_conformedit(request):
    '''确认编辑奖金信息
    V1.1不删除全部，不删除主表，只删除从表。'''
    res = {'code': '200'}
    uid = request.POST.get('edit_uid')
    hh = request.FILES.get('bonus_pdf')  # 获得合同的变量,代表用户传的合同文件
    # bonus_name = request.POST.get('bonus_name')  # 代表字符串，只是为了拿名字传给后端用
    # temp = bonus_name.split('\n')  # 对span标签的值做一个提取
    input_data = request.POST.get('inputdata')

    newdata = util.formdata_to_dict(input_data)
    _dic_start = util.dic_slice(newdata, 'nature', 'sale')
    _dic_end = util.dic_slice(newdata, 'MAX_FILE_SIZE', 'remarks')
    dd = {}
    dd.update(_dic_start)
    dd.update(_dic_end)
    ss = {}
    for i, j in newdata.items():
        if not i in dd:
            ss[i] = j
    new_list = util.dict_chunk(ss, 3)
    user = request.user
    BonusSystem.objects.filter(bonusunit_id=uid).delete()  # 删除所有从表
    # try:

    try:
        util.Autonomous_updateition(new_list, dd, user, hh, uid)
        res['msg'] = '更新成功'
    except:
        res['error'] = '错误，可能数据库中已存在该数据或者您提交的数据类型为空'

    return JsonResponse(res, safe=False)


def bonus_conformdel(request):
    res = {'code': '200'}
    delete_uid = request.GET.get('delete_uid')
    try:
        BonusUnit.objects.filter(id=delete_uid).delete()
        res['msg'] = '成功'
    except:
        res['error'] = '失败'

    return JsonResponse(res, safe=False)


def query_date(request):
    user = request.user
    nature1 = request.GET.get('nature1', '')
    pm1 = request.GET.get('pm1', '')
    sale1 = request.GET.get('sale1', '')
    unit1 = request.GET.get('unit1', '')
    # bonus1 = request.GET.get('bonus1', '')
    # ebonus = util.encrypt(bonus1)  # 加密前端传来的奖金信息传给后端进行匹配
    level1 = request.GET.get('level1', '')
    datefilter = request.GET.get('datefilter', '')
    query_dict = {}
    if nature1:
        query_dict['nature'] = nature1
    if pm1:
        query_dict['pm'] = pm1
    if sale1:
        query_dict['sale'] = sale1
    if unit1:
        query_dict['unit__contains'] = unit1
    # if bonus1:
    #     query_dict['bonus'] = ebonus
    if level1:
        if level1 == '二级':
            levelquery = [1, 2, 3]
            query_dict['bonussystem__level__in'] = levelquery
        if level1 == '三级':
            levelquery = [4, 5, 6]
            query_dict['bonussystem__level__in'] = levelquery
    if datefilter:
        times = Convert_time_format(datefilter)
        start_date = datetime.date(int(times[0][2]), int(times[0][0]), int(times[0][1]))
        end_date = datetime.date(int(times[1][2]), int(times[1][0]), int(times[1][1]))

        query_dict['bonussystem__completion_time__range'] = (start_date, end_date)
    bonuslist = BonusUnit.objects.filter(user=user).filter(**query_dict)
    return bonuslist, query_dict, datefilter


def Convert_time_format(timestr):
    s = timestr.split('- ')
    times = []
    for i in s:
        i = i.split('/')
        times.append(i)
    return times


def UnitView(request):
    '''单位筛选'''
    user = request.user
    queryset = BonusUnit.objects.filter(user=user)
    for field in BonusUnit._meta.fields:
        params = request.GET.get(field.name, None)
        if params:
            params = params.split(',')
            filter_fields = [Q(**{field.name: param}) for param in params]
            filter_field = reduce(lambda x, y: x | y, filter_fields)
            queryset = queryset.filter(filter_field)
    return queryset


def download_pdf(request):
    '''下载pdf文件'''
    dow_uid = request.GET.get('dow_uid')
    contract_pdf = BonusUnit.objects.filter(id=dow_uid).values('contract_pdf')
    contract_pdf = list(contract_pdf)
    contract_path = contract_pdf[0]['contract_pdf']
    filename = contract_path.split('/')[1]

    contract_path = os.path.join(settings.MEDIA_ROOT, contract_path)
    if not os.path.isfile(contract_path):
        return HttpResponse("Sorry but Not Found the File")

    def file_iterator(file_path, chunk_size=512):
        """
        文件生成器,防止文件过大，导致内存溢出
        :param file_path: 文件绝对路径
        :param chunk_size: 块大小
        :return: 生成器
        """

        with open(file_path, mode='rb') as f:
            while True:
                c = f.read(chunk_size)
                if c:
                    yield c
                else:
                    break

    try:

        response = StreamingHttpResponse(file_iterator(contract_path))
        # 以流的形式下载文件,这样可以实现任意格式的文件下载
        response['Content-Type'] = 'application/pdf'
        # Content-Disposition就是当用户想把请求所得的内容存为一个文件的时候提供一个默认的文件名
        response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(
            escape_uri_path(filename))  # 文件名可设置为中文
    except:
        log.error("Sorry but Not Found the File")
        return HttpResponse("Sorry but Not Found the File")

    return response
    # return HttpResponse('ok')


def output_excel(request):
    '''导出项目奖金到excel文件'''
    user = request.user

    bonuslist = BonusUnit.objects.filter(user=user).values('nature', 'unit', 'bonus', 'pm', 'sale',
                                                           'bonussystem__system', 'bonussystem__level',
                                                           'bonussystem__completion_time', 'remarks')
    for i in bonuslist:
        if i['bonus']:
            m = util.decrypt(i['bonus'])

            i['bonus'] = int(float(m))
        if i['pm']:
            pm = request.session.get(i['pm'])
            i['pm'] = pm
        if i['sale']:
            sale = request.session.get(i['sale'])
            i['sale'] = sale
        if i['bonussystem__level']:
            choices = (
                ('1', 'S2A2G2'), ('2', 'S2A1G2'), ('3', 'S1A2G2'), ('4', 'S3A3G3'), ('5', 'S3A2G3'), ('6', 'S2A3G3'))
            temp = i['bonussystem__level']
            print("temp", temp)
            level = choices[int(temp) - 1][1]
            i['bonussystem__level'] = level
    path = os.path.join(settings.MEDIA_ROOT, '奖金.xlsx')

    # 写入到Excel
    datas = list(bonuslist)

    write_to_excel(datas, path, sheetname='奖金.xlsx')

    with open(path, 'rb') as fr:
        content = fr.read()
    t = datetime.datetime.now().strftime('%Y-%m-%d')
    excel_name = '奖金表' + str(t) + '.xlsx'
    filename = urlquote(excel_name)

    response = HttpResponse(content)
    response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response['Content-Disposition'] = 'attachment; filename=%s' % (filename)
    os.remove(path)
    return response


def write_to_excel(data: list, path: str, sheetname):
    """
    把数据库写入excel
    :param data:
    :param path:
    :return:
    """
    # 实例化一个wrokbook
    wrokbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = wrokbook.active
    # 为sheet命名
    sheet.title = sheetname
    # 准备keys
    keys = data[0].keys()
    # 写标题的第一行
    get_title = ["单位性质", "单位名称", "奖金金额", "项目经理", "商务经理", '系统名称', '系统级别', '完成时间', '备注']
    for index, item in enumerate(get_title):
        sheet.cell(row=1, column=index + 1, value=item)

    # 准备写入数据
    for index, item in enumerate(data):
        # 遍历每个元素
        for k, v in enumerate(keys):
            # 这里要从第二列开始
            sheet.cell(row=index + 2, column=k + 1, value=str(item[v]))
    # 写入到文件
    wrokbook.save(path)
