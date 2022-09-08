import os
import re
import time
from django.utils.http import urlquote

import openpyxl

from django.views.decorators.csrf import csrf_exempt
from system.models import *
from crm import settings
from djcp.models import *
from system import models
from django.http import JsonResponse, HttpResponse, Http404, StreamingHttpResponse

from . import formdata, util


def get_unitinfo(request):
    '''拿到当前登录用户下的所有单位信息'''
    res = {"code": 100}
    user = request.user
    unitlist = Customer.objects.filter(user=user).order_by('-id').values('id', 'unit_name')
    unitlist = list(unitlist)

    return unitlist


def get_unittosys(unit_id):
    '''
    拿到单位对应下的系统信息
    :param unit_id: 单位的id
    :return:
    '''
    unittosyslist = ProInfo.objects.filter(proinfo_id=unit_id).all().order_by('-id').values('id', 'sys_name')
    unittosyslist = list(unittosyslist)
    return unittosyslist


def get_unit(request):
    '''提供给单位信息表中发送ajax'''
    res = {"code": 100}
    user = request.user
    unitlist = Customer.objects.filter(user=user).order_by('-id').values('id', 'unit_name')
    unitlist = list(unitlist)

    return JsonResponse(unitlist, safe=False)


def get_evrole():
    '''拿到角色名称为   测评师   的所有用户信息'''
    roleobj = Role.objects.filter(title__contains='测评师').values_list('evaluate__name', 'evaluate__id')

    evalist = []
    for i in roleobj:
        if i != (None, None):
            evalist.append(i)

    return evalist


def get_unit_id(request):
    '''拿到单位列表'''
    unit_id = request.GET.get('unit_id')
    unittosyslist = get_unittosys(unit_id)
    return JsonResponse(unittosyslist, safe=False)


def get_contact_id(request):
    '''选择单位信息自动填充联系人信息'''
    unit_id = request.GET.get('unit_id')
    contactlist = CUserInfo.objects.filter(nameinfo_id=unit_id).all().order_by('-id').values('id', 'name')
    contactlist = list(contactlist)
    return JsonResponse(contactlist, safe=False)


def get_gmanage_id(request):
    '''选择单位信息自动填充联系人信息'''
    unit_id = request.GET.get('unit_id')
    gmanagelist = GmanagerInfo.objects.filter(gnameinfo_id=unit_id).all().order_by('-id').values('id', 'gmanager')
    gmanagelist = list(gmanagelist)
    return JsonResponse(gmanagelist, safe=False)


def get_date_id(request):
    '''选择单位信息自动填充日期信息'''
    unit_id = request.GET.get('unit_id')
    datelist = ProDate.objects.filter(unit_id=unit_id).all().order_by('-id').values('id', 'contract')
    datelist = list(datelist)

    return JsonResponse(datelist, safe=False)


class AutoSerialNumber(object):
    """创建OA单号"""

    def __init__(self):
        # J201906120001
        # self.fd_apply_no = ApplicationBasicFormModel.delete_objects.filter(fd_apply_no__contains="J").order_by(
        #     "-fd_apply_no").first().fd_apply_no
        self.fd_apply_no = "J20196120001"
        self.date_str = self.fd_apply_no[1: 9]  # 日期字符串
        self._serial_number = self.fd_apply_no[9:]  # 流水号字符串
        self._serial_number = 0  # 流水号

    @property
    def serial_number(self):
        return self._serial_number

    @serial_number.setter
    def serial_number(self, value):
        if isinstance(value, int):
            self._serial_number = value
        else:
            self._serial_number = 1

    def __iter__(self):
        return self

    def __next__(self):
        self.serial_number += 1
        # 生成一个固定4位数的流水号
        return "{0:03d}".format(self.serial_number)

    def __call__(self, *args, **kwargs):
        # 返回生成序列号(日期加流水号)
        return "J" + self.date_str + next(self)

    # 时间格式化,最好是用定时器来调用该方法
    def timed_clear_serial_number(self):
        """用于每天定时清除流水号"""

        self.serial_number = 1
        self.date_str = time.strftime("%Y%m%d", time.localtime(time.time()))


def selectcontact(request):
    '''点击查询单位联系人信息'''
    user = request.user
    contactlist = CUserInfo.objects.filter(nameinfo__user=user)
    return contactlist


def selectgmanage(request):
    '''点击查询单位联系人信息'''
    user = request.user
    gmanagelist = GmanagerInfo.objects.filter(gnameinfo__user=user)
    return gmanagelist


def editcontact(request):
    '''点击编辑按钮事件拿到单位联系人的信息'''
    res = {'code': "200"}
    contact_edit = request.GET.get('contact_edit')
    contact = CUserInfo.objects.filter(id=contact_edit).values('name', 'email', 'post', 'phone', 'telephone',
                                                               'nameinfo')
    if contact:
        print(contact)
        contact = list(contact)
        res['contact'] = contact
    else:
        res['error'] = '无此联系人信息'
    return JsonResponse(res, safe=False)


@csrf_exempt
def conformeditcontact(request):
    '''确认编辑联系人信息'''
    res = {'code': '200'}

    if request.is_ajax():
        res = {'user': None, 'msg': None}
        userobj = formdata.UserInfoForm(request.POST)
        if userobj.is_valid():
            name = userobj.cleaned_data.get('name')
            email = userobj.cleaned_data.get('email')
            post = userobj.cleaned_data.get('post')
            phone = userobj.cleaned_data.get('phone')
            telephone = userobj.cleaned_data.get('telephone')
            nameinfo_id = request.POST.get('bind_unitcontact')
            uuid = request.POST.get('uuid')

            context = {"name": name, "email": email, "post": post, "phone": phone, "telephone": telephone,
                       "nameinfo_id": nameinfo_id}

            user = CUserInfo.objects.filter(id=uuid).update(**context)

            res['msg'] = '更新成功'
        else:
            res['error'] = userobj.errors

        return JsonResponse(res, safe=False)


def conformdelcontact(request):
    '''联系人确认删除的处理函数'''
    res = {'code': '200'}
    delete_uid = request.GET.get('delete_uid')
    try:
        CUserInfo.objects.filter(id=delete_uid).delete()
        res['msg'] = '成功'
    except:
        res['error'] = '失败'

    return JsonResponse(res, safe=False)


def editgmanager(request):
    '''点击编辑按钮事件拿到单位负责的信息'''
    res = {'code': "200"}
    gmanager_edit = request.GET.get('gmanager_edit')

    gmanager = GmanagerInfo.objects.filter(id=gmanager_edit).values('gmanager', 'gemail', 'gpost', 'gphone',
                                                                    'gtelephone',
                                                                    'gnameinfo')
    if gmanager:

        gmanager = list(gmanager)
        res['gmanager'] = gmanager
    else:
        res['error'] = '无此联系人信息'
    return JsonResponse(res, safe=False)


@csrf_exempt
def conformeditgmanager(request):
    '''确认编辑负责人人信息'''
    res = {'code': '200'}

    if request.is_ajax():
        res = {'user': None, 'msg': None}
        guserobj = formdata.GmanagerInfoForm(request.POST)
        if guserobj.is_valid():
            gname = guserobj.cleaned_data.get('gname')
            gemail = guserobj.cleaned_data.get('gemail')
            gpost = guserobj.cleaned_data.get('gpost')
            gphone = guserobj.cleaned_data.get('gphone')
            gtelephone = guserobj.cleaned_data.get('gtelephone')
            gnameinfo_id = request.POST.get('bind_unitgmanager')
            uuid = request.POST.get('uuid')

            context = {"gmanager": gname, "gemail": gemail, "gpost": gpost, "gphone": gphone, "gtelephone": gtelephone,
                       "gnameinfo_id": gnameinfo_id}

            GmanagerInfo.objects.filter(id=uuid).update(**context)

            res['msg'] = '更新成功'
        else:
            res['error'] = guserobj.errors

        return JsonResponse(res, safe=False)


def conformdelgmanager(request):
    '''联系人确认删除的处理函数'''
    res = {'code': '200'}
    delete_uid = request.GET.get('delete_uid')
    try:
        GmanagerInfo.objects.filter(id=delete_uid).delete()
        res['msg'] = '成功'
    except:
        res['error'] = '失败'

    return JsonResponse(res, safe=False)


def editunit(request):
    '''点击编辑按钮事件拿到被测单位的信息'''
    res = {'code': "200"}
    unit_edit = request.GET.get('unit_edit')

    unit = Customer.objects.filter(id=unit_edit).values('unit_name', 'company_profile', 'address', 'nature',
                                                        'code', 'department', 'superdepar')
    if unit:

        unit = list(unit)
        res['unit'] = unit
    else:
        res['error'] = '无此单位信息'
    return JsonResponse(res, safe=False)


@csrf_exempt
def conformeditunit(request):
    '''确认编辑负责人人信息'''
    res = {'code': '200'}

    res = {'user': None, 'msg': None}
    uuid = request.POST.get('uuid')
    unit_name = request.POST.get('unit_name', '')
    address = request.POST.get('address', '')
    nature = request.POST.get('nature', '')
    code = request.POST.get('code', '')
    department = request.POST.get('department', '')
    superdepartment = request.POST.get('superdepartment', '')
    desc = request.POST.get('desc', '')
    try:
        Customer.objects.filter(id=uuid).update(
            unit_name=unit_name, address=address, nature=nature, code=code,
            department=department, superdepar=superdepartment,
            company_profile=desc)
        res['msg'] = '更新成功'
    except:
        res['error'] = '更新失败'

    return JsonResponse(res, safe=False)


def conformdelunit(request):
    '''联系人确认删除的处理函数'''
    res = {'code': '200'}
    delete_uid = request.GET.get('delete_uid')
    try:
        Customer.objects.filter(id=delete_uid).delete()
        res['msg'] = '成功'
    except:
        res['error'] = '失败'

    return JsonResponse(res, safe=False)


def editdate(request):
    '''点击编辑按钮事件拿到单位日期的信息'''
    res = {'code': "200"}
    date_edit = request.GET.get('date_edit')

    date = ProDate.objects.filter(id=date_edit).values('apply', 'accept', 'formalization', 'contract', 'secret',
                                                       'authorization',
                                                       'task', 'question', 'scheme', 'schemereview',
                                                       'firstmeeting', 'onsiteevalaation', 'record', 'test',
                                                       'lasttmeeting',
                                                       'reporter_date', 'reporterview', 'acceptreporter', 'agreen',
                                                       'unit_id', )
    if date:

        date = list(date)
        res['date'] = date
    else:
        res['error'] = '无此日期信息'
    return JsonResponse(res, safe=False)


@csrf_exempt
def conformeditdate(request):
    '''确认编辑时间信息'''
    res = {'code': '200'}

    if request.is_ajax():
        unit_id = request.POST.get('bind_unitcontact')
        if unit_id == '0':
            res['error'] = '未绑定单位名称'
            return JsonResponse(res)
        uuid = request.POST.get('uuid')
        apply = request.POST.get('apply')
        accept = request.POST.get('accept')
        formalization = request.POST.get('formalization')
        contract = request.POST.get('contract')
        secret = request.POST.get('secret')
        authorization = request.POST.get('authorization')

        task = request.POST.get('task')
        question = request.POST.get('question')
        scheme = request.POST.get('scheme')
        schemereview = request.POST.get('schemereview')

        firstmeeting = request.POST.get('firstmeeting')
        onsiteevalaation = request.POST.get('onsiteevalaation')
        record = request.POST.get('record')
        test = request.POST.get('test')
        lasttmeeting = request.POST.get('lasttmeeting')

        reporter = request.POST.get('reporter')
        reporterview = request.POST.get('reporterview')
        acceptreporter = request.POST.get('acceptreporter')
        agreen = request.POST.get('agreen')

        prodate = {"apply": apply, "accept": accept, "formalization": formalization, "contract": contract,
                   "secret": secret, "authorization": authorization,
                   "task": task, "question": question, "scheme": scheme,
                   "schemereview": schemereview, "firstmeeting": firstmeeting, "onsiteevalaation": onsiteevalaation,
                   "record": record,
                   "test": test,
                   "lasttmeeting": lasttmeeting, "reporter_date": reporter, "reporterview": reporterview,
                   "acceptreporter": acceptreporter, "agreen": agreen,
                   "unit_id": unit_id, 'id': uuid
                   }
        for k, v in prodate.items():
            if v == '':
                res['error'] = '{}时间未选择'.format(k)
                return JsonResponse(res)

        ProDate.objects.filter(id=uuid).update(**prodate)

    return JsonResponse(res, safe=False)


def conformdeldate(request):
    '''联系人确认删除的处理函数'''
    res = {'code': '200'}
    delete_uid = request.GET.get('delete_uid')
    try:
        ProDate.objects.filter(id=delete_uid).delete()
        res['msg'] = '成功'
    except:
        res['error'] = '失败'

    return JsonResponse(res, safe=False)


def editsystem(request):
    '''点击编辑按钮事件拿到单位系统的信息'''
    res = {'code': "200"}
    system_edit = request.GET.get('system_edit')

    system = ProInfo.objects.filter(id=system_edit).values('proname', 'agentname', 'sys_name', 'level', 'sys_service',
                                                           'sys_obj', 'pro_num', 'Record_num', 'pm', 'supervisor',
                                                           'reporter_pro', 'supervisored', 'desc', 'proinfo_id')
    if system:

        system = list(system)
        res['system'] = system
    else:
        res['error'] = '无此系统信息'
    return JsonResponse(res, safe=False)


@csrf_exempt
def conformeditsystem(request):
    '''确认编辑系统信息'''
    res = {'code': '200'}
    uuid = request.POST.get("uuid")
    proname = request.POST.get("proname")
    agentname = request.POST.get("agentname", '')
    sys_name = request.POST.get("sys_name")
    level = request.POST.get("level", '')
    sys_service = request.POST.get("sys_service", '')
    sys_obj = request.POST.get("sys_obj", '')
    pro_num = request.POST.get("pro_num", '')
    Record_num = request.POST.get("Record_num", '')
    pm = request.POST.get("pm", '')
    supervisor = request.POST.get("supervisor", '')
    reporter = request.POST.get("reporter", '')
    supervisored = request.POST.getlist("supervisored")
    desc = request.POST.get("desc", '')
    proinfo_id = request.POST.get('bind_systeminfo_name')

    pro_dict = {"proname": proname, "sys_obj": sys_obj, "agentname": agentname, "sys_name": sys_name,
                "level": level, "sys_service": sys_service,
                "pro_num": pro_num, "Record_num": Record_num, "pm": pm,
                "supervisor": supervisor, "reporter_pro": reporter, "supervisored": supervisored, "desc": desc,
                "proinfo_id": proinfo_id}

    if proname == '':
        res['error'] = '项目名称不能为空'
    elif sys_name == '':
        res['error'] = '系统名称不能为空'
    elif desc == '':
        res['error'] = '描述不能为空'
    elif level == '0':
        res['error'] = '未选择系统级别'
    elif sys_service == '0':
        res['error'] = '未选择系统服务性质'
    elif sys_obj == '0':
        res['error'] = '未选择系统服务对象'
    elif pm == '0':
        res['error'] = '未选择项目经理'
    elif supervisor == '0':
        res['error'] = '未选择监督人'
    elif reporter == '0':
        res['error'] = '未选择汇报人'
    elif proinfo_id == '0':
        res['error'] = '未绑定单位信息'
    elif supervisored == []:
        res['error'] = '未选择被监督人'
    else:

        ProInfo.objects.filter(id=uuid).update(**pro_dict)
        res['msg'] = '更新成功'

    return JsonResponse(res, safe=False)


def conformdelsystem(request):
    '''联系人确认删除的处理函数'''
    res = {'code': '200'}
    delete_uid = request.GET.get('delete_uid')
    try:
        ProInfo.objects.filter(id=delete_uid).delete()
        res['msg'] = '成功'
    except:
        res['error'] = '失败'

    return JsonResponse(res, safe=False)


@csrf_exempt
def addproject(request):
    res = {'user': None, 'msg': None}

    customer = request.POST.get('bind_unitcontact')

    proInfo = request.POST.get('bind_sys')
    prodate = request.POST.get('condate')
    cuserInfo = request.POST.get('contact')
    gmanagerInfo = request.POST.get('gmanage')
    remark = request.POST.get('remark', '')

    try:
        Djcp.objects.create(remark=remark, customer_id=customer, proInfo_id=proInfo, prodate_id=prodate,
                            cuserInfo_id=cuserInfo,
                            gmanagerInfo_id=gmanagerInfo)
        res['msg'] = '添加成功'
        return JsonResponse(res, safe=False)
    except:
        res['error'] = '不知名的错误'
        return JsonResponse(res, safe=False)


def conformdelproject(request):
    '''联系人确认删除的处理函数'''
    res = {'code': '200'}
    delete_uid = request.GET.get('delete_uid')
    try:
        Djcp.objects.filter(id=delete_uid).delete()
        res['msg'] = '成功'
    except:
        res['error'] = '失败'

    return JsonResponse(res, safe=False)


def write_to_excel(data: list, path: str, sheetname, remaindlist: list):
    """
    把数据库写入excel
    :param data:
    :param path:
    :return:
    """
    # 实例化一个wrokbook
    wrokbook = openpyxl.Workbook()
    # 激活一个sheet
    sheet = wrokbook.active
    # 为sheet命名
    sheet.title = sheetname
    # 准备keys
    keys = data[0].keys()
    # 写标题的第一行
    get_title = Customer.get_title_list() + CUserInfo.get_title_list() + GmanagerInfo.get_title_list() + ProInfo.get_title_list() + ProDate.get_title_list() + remaindlist
    for index, item in enumerate(get_title):
        sheet.cell(row=1, column=index + 1, value=item)

    # 准备写入数据
    for index, item in enumerate(data):
        # 遍历每个元素
        for k, v in enumerate(keys):
            # 这里要从第二列开始
            sheet.cell(row=index + 2, column=k + 1, value=str(item[v]))
    # 写入到文件
    wrokbook.save(path)


def merge_dicts(*dict_args):
    '''定义多个字典合并的函数'''
    result = {}
    for item in dict_args:
        result.update(item)
    return result


def download(request):
    '''点击下载按钮弹出下载对话框的试图函数'''
    res = {'code': "200"}
    dow_uid = request.GET.get('dow_uid')

    project = Djcp.objects.filter(id=dow_uid).values('customer', 'proInfo', 'prodate', 'cuserInfo', 'gmanagerInfo',
                                                     'remark')
    if project:

        project = list(project)
        res['project'] = project
    else:
        res['error'] = '无此日期信息'
    return JsonResponse(res, safe=False)


def conformdownload(request):
    """
           导出数据到excel
           :param request:
           :return:
           """
    res = {'code': '200'}

    # 准备写入的路径
    dow_uid = request.GET.get('dow_uid')
    dow_uid = int(dow_uid)

    proobj = Djcp.objects.filter(id=dow_uid).first()
    customer_id = proobj.customer_id
    cuserinfo_id = proobj.cuserInfo_id

    proinfo_id = proobj.proInfo_id
    prodate_id = proobj.prodate_id
    gmanager_id = proobj.gmanagerInfo_id

    cus = Customer.objects.filter(id=customer_id).values('unit_name', 'company_profile', 'address', 'nature', 'code',
                                                         'department',
                                                         'superdepar')
    cuserinfo = CUserInfo.objects.filter(id=cuserinfo_id).values('name', 'email', 'post', 'phone', 'telephone')
    gmanager = GmanagerInfo.objects.filter(id=gmanager_id).values('gmanager', 'gphone', 'gtelephone', 'gpost', 'gemail')
    proinfo = ProInfo.objects.filter(id=proinfo_id).values('proname', 'agentname', 'sys_name', 'level', 'sys_service',
                                                           'sys_obj', 'pro_num', 'Record_num', 'pm', 'supervisor',
                                                           'reporter_pro', 'supervisored', 'desc')
    prodate = ProDate.objects.filter(id=prodate_id).values('apply', 'accept', 'formalization', 'contract', 'secret',
                                                           'authorization', 'task', 'question', 'scheme',
                                                           'schemereview', 'firstmeeting', 'onsiteevalaation', 'record'
                                                           , 'test', 'lasttmeeting', 'reporter_date', 'reporterview',
                                                           'acceptreporter', 'agreen')

    datas = []  # 存放处理后的干净数据的最终传给excel
    remainddate = {}  # 存放剩下14个字段的

    cus = list(cus)
    cuserinfo = list(cuserinfo)
    gmanager = list(gmanager)
    proinfo = list(proinfo)
    prodate = list(prodate)

    cus = cus[0]
    cuserinfo = cuserinfo[0]
    gmanager = gmanager[0]
    proinfo = proinfo[0]
    prodate = prodate[0]

    '''处理项目信息的函数'''
    for i, j in proinfo.items():
        if i == "level":
            eva = ProInfo.objects.filter(id=proinfo_id).first()
            m = eva.get_level_display()
            proinfo[i] = m
            tt = {
                '1': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ 🗹第二级  A_2_ S _2_ □第三级 A__ S __',
                '2': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ 🗹第二级  A_1_ S _2_ □第三级 A__ S __',
                '3': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ 🗹第二级  A_2_ S_1_  □第三级 A__ S __',
                '4': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ □第二级  A__ S __   🗹第三级 A_3_ S _3_',
                '5': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ □第二级  A__ S __   🗹第三级 A_2_ S _3_',
                '6': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ □第二级  A__ S __   🗹第三级 A_3_ S _2_',
            }

            remainddate['formal_level'] = tt.get(j)  # 形式化等级

            hh = {
                "1": '''□第一级 A___  S___          🗹第二级 A_2_ S _2_\n□第三级 A__  S__         □第四级  A__ S __ ''',
                "2": '''□第一级 A___  S___          🗹第二级 A_1_ S _2_\n□第三级 A__  S__         □第四级  A__ S __ ''',
                "3": '''□第一级 A___  S___          🗹第二级 A_2_ S _1_\n□第三级 A__  S__         □第四级  A__ S __ ''',
                "4": '''□第一级 A___  S___          □第二级 A__ S __\n🗹第三级 A_3_  S_3_         □第四级  A__ S __ ''',
                "5": '''□第一级 A___  S___          □第二级 A__ S __\n🗹第三级 A_2_  S_3_         □第四级  A__ S __ ''',
                "6": '''□第一级 A___  S___          □第二级 A__ S __\n🗹第三级 A_3_  S_2_         □第四级  A__ S __ ''',
            }
            remainddate['sspl'] = hh.get(j)  # 系统安全保护等级
            mm = re.findall(r"\d+\.?\d*", m)  # 提取S2A2G2中的数字

            aalist = []
            for i in mm:
                i = int(i)
                i = util.num_to_char(i)
                aalist.append(i)

            remainddate['slevel'] = aalist[0]  # 业务信息安全保护等级
            remainddate['alevel'] = aalist[1]  # 系统服务安全保护等级
            remainddate['glevel'] = aalist[2]  # 安全保护等级

        if i == "sys_service":
            sys_servicedict = {
                '1': '🗹业务专网      ☐互联网       ☐其它',
                '2': '☐业务专网      🗹互联网       ☐其它',
                '3': '☐业务专网      ☐互联网       🗹其它',
            }

            proinfo[i] = sys_servicedict.get(j)
        if i == "sys_obj":
            sys_objedict = {
                '1': '🗹 单位内部人员  ☐ 社会公众人员 ☐两者均包括  ☐其他',
                '2': '☐ 单位内部人员  🗹 社会公众人员 ☐两者均包括  ☐其他',
                '3': '☐ 单位内部人员  ☐ 社会公众人员 🗹两者均包括  ☐其他',
                '4': '☐ 单位内部人员  ☐ 社会公众人员 ☐两者均包括  🗹 其他',
            }

            proinfo[i] = sys_objedict.get(j)

        if i == 'supervisor':
            supervisor = int(j)
            eva = models.Evaluate.objects.filter(id=supervisor).values()[0]
            j = eva.get('name')
            proinfo[i] = j

        if i == "pm":
            pm = int(j)
            eva = models.Evaluate.objects.filter(id=pm).values()[0]
            j = eva.get('name')
            evaphone = eva.get('phone')
            remainddate['preparer'] = j  # 1:填表人数据
            remainddate['evaphone'] = evaphone  # 1:填表人电话
            proinfo[i] = j

        if i == "reporter_pro":
            reporter_pro = int(j)
            eva = models.Evaluate.objects.filter(id=reporter_pro).values()[0]
            j = eva.get('name')
            proinfo[i] = j
        if i == "supervisored":
            supervisored = []
            m = re.findall(r"\d+\.?\d*", j)
            for j in m:
                eva = models.Evaluate.objects.filter(id=j).values()[0]
                supervisored.append(eva.get('name'))
            supervisored = str(supervisored)
            supervisored = re.sub("'", '', supervisored)
            supervisored = supervisored.strip('[]')
            supervisored = supervisored.replace(',', '、')
            proinfo[i] = supervisored
        if i == "sys_name":
            remainddate['recordtable'] = j + '备案表'  # 备案表
            remainddate['letable'] = j + '定级报告'  # 定级报告
            remainddate['leaddtable'] = j + '定级备案补充信息表'  # 定级备案补充信息表
    '''处理单位信息的函数'''
    for i, j in cus.items():
        if i == "nature":
            naturedict = {
                '1': '🗹党政机关      ☐国家重要行业、重要领域或重要企事业单位       ☐一般企事业单位       ☐其他类型',
                '2': '☐党政机关      🗹国家重要行业、重要领域或重要企事业单位       ☐一般企事业单位       ☐其他类型',
                '3': '☐党政机关      ☐国家重要行业、重要领域或重要企事业单位       🗹一般企事业单位       ☐其他类型',
                '4': '☐党政机关      ☐国家重要行业、重要领域或重要企事业单位       ☐一般企事业单位       🗹其他类型',
            }
            cus[i] = naturedict.get(j)
    '''处理日期格式的函数'''
    ff = []
    for i, j in prodate.items():
        if i == 'contract':
            contime = util.datetomonth(j)[:-3]
            ff.append(contime)
        if i == 'question':
            question = util.datetomonth(j)
            remainddate['questionface'] = question  # 信息调查表封面
        if i == 'test':
            test = util.datetomonth(j)
            remainddate['testmen'] = test  # 监督人签字时间
        if i == 'agreen':
            agreetime = util.datetomonth(j)[:-3]
            ff.append(agreetime)
        m = util.datetomonth(j)
        prodate[i] = m

    remainddate['servereva'] = ff[0] + '-' + ff[1]  # 服务评价表时间
    remainddate['acceptunit'] = '上海交通大学信息安全服务技术研究实验室'  # 接收单位

    '''处理剩余的14个字段的函数'''

    project1 = merge_dicts(cus, cuserinfo, gmanager, proinfo, prodate, remainddate)

    datas.append(project1)

    sheetname = proinfo.get("proname") + str(dow_uid)
    excel_name = sheetname + '.xlsx'

    path = os.path.join(settings.MEDIA_ROOT, excel_name)

    # 写入到Excel
    try:
        remaindlist = ["{备案表名称}", "{定级报告名称}", "{补充表名称}", "{形式化等级}", "{系统安全保护等级}", "{业务信息安全等级}", "{系统服务安全等级}",
                       "{安全保护等级}",
                       "{填表人}", "{项目负责人联系电话}", "{信息调查表封面}", "{监督人签字时间}", "{服务情况评价表时间}", "{接收单位}"]
        write_to_excel(datas, path, sheetname=sheetname, remaindlist=remaindlist)

        res['msg'] = '生成excel成功'
        return JsonResponse(res)
    except:
        res['error'] = '遇到不知名的错误'
        return JsonResponse(res)


def stream_http_download(request, file_path):
    try:
        response = StreamingHttpResponse(open(file_path, 'rb'))
        response['content_type'] = "application/octet-stream"
        response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(file_path)
        return response
    except Exception:
        raise Http404


def download1(request):
    '''下载的函数'''

    res = {'code': '200'}

    # 准备写入的路径
    dow_uid = request.GET.get('dow_uid')
    dow_uid = int(dow_uid)

    proobj = Djcp.objects.filter(id=dow_uid).first()
    customer_id = proobj.customer_id
    cuserinfo_id = proobj.cuserInfo_id

    proinfo_id = proobj.proInfo_id
    prodate_id = proobj.prodate_id
    gmanager_id = proobj.gmanagerInfo_id

    cus = Customer.objects.filter(id=customer_id).values('unit_name', 'company_profile', 'address', 'nature', 'code',
                                                         'department', 'superdepar')
    cuserinfo = CUserInfo.objects.filter(id=cuserinfo_id).values('name', 'email', 'post', 'phone', 'telephone')
    gmanager = GmanagerInfo.objects.filter(id=gmanager_id).values('gmanager', 'gphone', 'gtelephone', 'gpost', 'gemail')
    proinfo = ProInfo.objects.filter(id=proinfo_id).values('proname', 'agentname', 'sys_name', 'level', 'sys_service',
                                                           'sys_obj', 'pro_num', 'Record_num', 'pm', 'supervisor',
                                                           'reporter_pro', 'supervisored', 'desc')
    prodate = ProDate.objects.filter(id=prodate_id).values('apply', 'accept', 'formalization', 'contract', 'secret',
                                                           'authorization', 'task', 'question', 'scheme',
                                                           'schemereview', 'firstmeeting', 'onsiteevalaation', 'record'
                                                           , 'test', 'lasttmeeting', 'reporter_date', 'reporterview',
                                                           'acceptreporter', 'agreen')

    datas = []  # 存放处理后的干净数据的最终传给excel
    remainddate = {}  # 存放剩下14个字段的

    cus = list(cus)
    cuserinfo = list(cuserinfo)
    gmanager = list(gmanager)
    proinfo = list(proinfo)
    prodate = list(prodate)

    cus = cus[0]
    cuserinfo = cuserinfo[0]
    gmanager = gmanager[0]
    proinfo = proinfo[0]
    prodate = prodate[0]

    '''处理项目信息的函数'''
    for i, j in proinfo.items():
        if i == "level":
            eva = ProInfo.objects.filter(id=proinfo_id).first()
            m = eva.get_level_display()
            proinfo[i] = m
            tt = {
                '1': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ 🗹第二级  A_2_ S _2_ □第三级 A__ S __',
                '2': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ 🗹第二级  A_1_ S _2_ □第三级 A__ S __',
                '3': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ 🗹第二级  A_2_ S_1_  □第三级 A__ S __',
                '4': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ □第二级  A__ S __   🗹第三级 A_3_ S _3_',
                '5': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ □第二级  A__ S __   🗹第三级 A_2_ S _3_',
                '6': '□ 不确定           🗹确定，安全保护等级是：\n□第一级 A___  S___ □第二级  A__ S __   🗹第三级 A_3_ S _2_',
            }

            remainddate['formal_level'] = tt.get(j)  # 形式化等级

            hh = {
                "1": '''□第一级 A___  S___          🗹第二级 A_2_ S _2_\n□第三级 A__  S__         □第四级  A__ S __ ''',
                "2": '''□第一级 A___  S___          🗹第二级 A_1_ S _2_\n□第三级 A__  S__         □第四级  A__ S __ ''',
                "3": '''□第一级 A___  S___          🗹第二级 A_2_ S _1_\n□第三级 A__  S__         □第四级  A__ S __ ''',
                "4": '''□第一级 A___  S___          □第二级 A__ S __\n🗹第三级 A_3_  S_3_         □第四级  A__ S __ ''',
                "5": '''□第一级 A___  S___          □第二级 A__ S __\n🗹第三级 A_2_  S_3_         □第四级  A__ S __ ''',
                "6": '''□第一级 A___  S___          □第二级 A__ S __\n🗹第三级 A_3_  S_2_         □第四级  A__ S __ ''',
            }
            remainddate['sspl'] = hh.get(j)  # 系统安全保护等级
            mm = re.findall(r"\d+\.?\d*", m)  # 提取S2A2G2中的数字

            aalist = []
            for i in mm:
                i = int(i)
                i = util.num_to_char(i)
                aalist.append(i)

            remainddate['slevel'] = aalist[0]  # 业务信息安全保护等级
            remainddate['alevel'] = aalist[1]  # 系统服务安全保护等级
            remainddate['glevel'] = aalist[2]  # 安全保护等级

        if i == "sys_service":
            sys_servicedict = {
                '1': '🗹业务专网      ☐互联网       ☐其它',
                '2': '☐业务专网      🗹互联网       ☐其它',
                '3': '☐业务专网      ☐互联网       🗹其它',
            }

            proinfo[i] = sys_servicedict.get(j)
        if i == "sys_obj":
            sys_objedict = {
                '1': '🗹 单位内部人员  ☐ 社会公众人员 ☐两者均包括  ☐其他',
                '2': '☐ 单位内部人员  🗹 社会公众人员 ☐两者均包括  ☐其他',
                '3': '☐ 单位内部人员  ☐ 社会公众人员 🗹两者均包括  ☐其他',
                '4': '☐ 单位内部人员  ☐ 社会公众人员 ☐两者均包括  🗹 其他',
            }

            proinfo[i] = sys_objedict.get(j)

        if i == 'supervisor':
            supervisor = int(j)
            eva = models.Evaluate.objects.filter(id=supervisor).values()[0]
            j = eva.get('name')
            proinfo[i] = j

        if i == "pm":
            pm = int(j)
            eva = models.Evaluate.objects.filter(id=pm).values()[0]
            j = eva.get('name')
            evaphone = eva.get('phone')
            remainddate['preparer'] = j  # 1:填表人数据
            remainddate['evaphone'] = evaphone  # 1:填表人电话
            proinfo[i] = j

        if i == "reporter_pro":
            reporter_pro = int(j)
            eva = models.Evaluate.objects.filter(id=reporter_pro).values()[0]
            j = eva.get('name')
            proinfo[i] = j
        if i == "supervisored":
            supervisored = []
            m = re.findall(r"\d+\.?\d*", j)
            for j in m:
                eva = models.Evaluate.objects.filter(id=j).values()[0]
                supervisored.append(eva.get('name'))
            supervisored = str(supervisored)
            supervisored = re.sub("'", '', supervisored)
            supervisored = supervisored.strip('[]')
            supervisored = supervisored.replace(',', '、')
            proinfo[i] = supervisored
        if i == "sys_name":
            remainddate['recordtable'] = j + '备案表'  # 备案表
            remainddate['letable'] = j + '定级报告'  # 定级报告
            remainddate['leaddtable'] = j + '定级备案补充信息表'  # 定级备案补充信息表
    '''处理单位信息的函数'''
    for i, j in cus.items():
        if i == "nature":
            naturedict = {
                '1': '🗹党政机关      ☐国家重要行业、重要领域或重要企事业单位       ☐一般企事业单位       ☐其他类型',
                '2': '☐党政机关      🗹国家重要行业、重要领域或重要企事业单位       ☐一般企事业单位       ☐其他类型',
                '3': '☐党政机关      ☐国家重要行业、重要领域或重要企事业单位       🗹一般企事业单位       ☐其他类型',
                '4': '☐党政机关      ☐国家重要行业、重要领域或重要企事业单位       ☐一般企事业单位       🗹其他类型',
            }
            cus[i] = naturedict.get(j)
    '''处理日期格式的函数'''
    ff = []
    for i, j in prodate.items():
        if i == 'contract':
            contime = util.datetomonth(j)[:-3]
            ff.append(contime)
        if i == 'question':
            question = util.datetomonth(j)
            remainddate['questionface'] = question  # 信息调查表封面
        if i == 'test':
            test = util.datetomonth(j)
            remainddate['testmen'] = test  # 监督人签字时间
        if i == 'agreen':
            agreetime = util.datetomonth(j)[:-3]
            ff.append(agreetime)
        m = util.datetomonth(j)
        prodate[i] = m

    remainddate['servereva'] = ff[0] + '-' + ff[1]  # 服务评价表时间
    remainddate['acceptunit'] = '上海交通大学信息安全服务技术研究实验室'  # 接收单位

    '''处理剩余的14个字段的函数'''

    project1 = merge_dicts(cus, cuserinfo, gmanager, proinfo, prodate, remainddate)

    datas.append(project1)

    sheetname = proinfo.get("proname")
    excel_name = sheetname + str(time.time()) + '.xlsx'

    path = os.path.join(settings.MEDIA_ROOT, excel_name)

    # 写入到Excel
    remaindlist = ["{备案表名称}", "{定级报告名称}", "{补充表名称}", "{形式化等级}", "{系统安全保护等级}", "{业务信息安全等级}", "{系统服务安全等级}", "{安全保护等级}",
                   "{填表人}", "{项目负责人联系电话}", "{信息调查表封面}", "{监督人签字时间}", "{服务情况评价表时间}", "{接收单位}"]

    write_to_excel(datas, path, sheetname=sheetname, remaindlist=remaindlist)

    with open(path, 'rb') as fr:
        content = fr.read()

    filename = urlquote(excel_name)
    response = HttpResponse(content)
    response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response['Content-Disposition'] = 'attachment; filename=%s' % (filename)
    os.remove(path)
    return response


def Download_assistant(request):
    '''下载小助手'''

    # 判断下载文件是否存在
    print(settings.DOWNLOAD_Helper)
    if not os.path.isfile(settings.DOWNLOAD_Helper):
        return HttpResponse("Sorry but Not Found the File")

    def file_iterator(file_path, chunk_size=512):
        """
        文件生成器,防止文件过大，导致内存溢出
        :param file_path: 文件绝对路径
        :param chunk_size: 块大小
        :return: 生成器
        """
        with open(file_path, mode='rb') as f:
            while True:
                c = f.read(chunk_size)
                if c:
                    yield c
                else:
                    break

    try:
        # 设置响应头
        # StreamingHttpResponse将文件内容进行流式传输，数据量大可以用这个方法（.pdf,.mp3,.mp4等等什么样格式的文件都可以下载）

        response = StreamingHttpResponse(file_iterator(settings.DOWNLOAD_Helper))
        # 以流的形式下载文件,这样可以实现任意格式的文件下载
        response['Content-Type'] = 'application/octet-stream'
        # Content-Disposition就是当用户想把请求所得的内容存为一个文件的时候提供一个默认的文件名
        response['Content-Disposition'] = 'attachment;filename={file_name}{format}'.format(
            file_name='Evaluation_assistant', format=".zip")
    except:
        return HttpResponse("Sorry but Not Found the File")

    # 在这里千万记得return,否则不会出现下载
    return response
